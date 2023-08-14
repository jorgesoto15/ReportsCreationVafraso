import os
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import NamedStyle, Font, Alignment
from typing import Union

class ExcelReport:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.names_to_search_jrz = []
        self.names_to_search_chih = []
        self.names_to_search_both = []
        self.rows_juarez = [1]
        self.rows_chih = [1]
        self.row_not_in_count = []
        self.currency_format = NamedStyle(name='currency', number_format='"$"#,##0.00')
        self.percentage_format = NamedStyle(name='percentage', number_format='0.00%')
    
    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)  # Load the report workbook
        except InvalidFileException:
            print("Error: Invalid file format. Openpyxl does not support this file format.")
            exit()
    
    def change_worksheet_name(self, old_sheet_name, new_sheet_name):
        worksheet = self.workbook[old_sheet_name]
        worksheet.title = new_sheet_name
        
    def delete_columns(self, sheet_name, columns_to_delete: Union[list, int]):
        worksheet = self.workbook[sheet_name]
        for col_index in sorted(columns_to_delete, reverse=True):
            worksheet.delete_cols(col_index)

    def delete_rows(self,sheet_name, column_index:int, values_to_delete:list):
        rows_to_delete = []
        worksheet = self.workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=2,min_col=column_index, max_col=column_index, values_only=False):
            cell_value = row[0].value
            if cell_value in values_to_delete:
                row_index = row[0].row
                rows_to_delete.append(row_index)

        for row_index in sorted(rows_to_delete, reverse= True):
            worksheet.delete_rows(row_index)
    
    def open_invoices(self, juarez_path, chih_path, both_path):
        try:
            with open(juarez_path, "r") as file:
                self.names_to_search_jrz = [name.strip() for name in file.readlines()]
            with open(chih_path, "r") as file:
                self.names_to_search_chih = [name.strip() for name in file.readlines()]
            with open(both_path, "r") as file:
                self.names_to_search_both = [name.strip() for name in file.readlines()]
        except FileNotFoundError:
            print("Error: Archivo de los nombres de las empresas no encontrado")
            exit()
    
    def filter_and_append_rows(self, worksheet_name):
        worksheet = self.workbook[worksheet_name]

        max_row = worksheet.max_row - 2

        try:
            for row in worksheet.iter_rows(min_row=2, max_row=max_row):
                cell_value = row[1].value
                cell_total = row[5].value
                if cell_value in self.names_to_search_jrz:
                     row_index = row[0].row
                     self.rows_juarez.append(row_index)
                elif cell_value in self.names_to_search_chih:
                    row_index_cuu = row[0].row
                    self.rows_chih.append(row_index_cuu)
                elif cell_value in self.names_to_search_both:
                    row_index = row[0].row
                    if cell_value == "NEFAB MEXICO":
                        if cell_total >= 100000:
                            self.rows_juarez.append(row_index)
                        else:
                            self.rows_chih.append(row_index)
                    else:
                        if 150000 >= cell_total >= 130000:
                            self.rows_juarez.append(row_index)
                        elif cell_total < 30000:
                            self.row_not_in_count.append(row_index)
                        else:
                            self.rows_chih.append(row_index)
                else:
                    print(str(cell_value) + " no está en el archivo de los nombres de las empresas")
                    exit()
        except:
            print("Error: Archivo de los nombres de las empresas no encontrado")
            exit()

    def create_new_worksheet(self, new_worksheet_name):
        self.workbook.create_sheet(new_worksheet_name)
    
    def change_column_width(self, columns_to_change: Union[list, str], new_width:float, specific_worksheets):
        specific_worksheets = [self.workbook[worksheet_name] for worksheet_name in specific_worksheets]
        
        for worksheet in specific_worksheets:
            for column in columns_to_change:
                worksheet.column_dimensions[column].width = new_width
         
    def copy_rows_to_ws(self, rows_to_copy: Union[list[int], int], worksheet_name, new_worksheet):
        maxc = self.workbook[worksheet_name].max_column

        for r in rows_to_copy:
            for c in range(1,maxc + 1):
                self.workbook[new_worksheet].cell(row=r, column=c).value = self.workbook[new_sheet_name].cell(row=r, column=c).value
    
    def delete_empty_rows(self, worksheet):
        index_row = []

        for i in range(1, self.workbook[worksheet].max_row):
            if self.workbook[worksheet].cell(i,1).value is None:
                index_row.append(i)
        
        for row_del in range(len(index_row)):
            self.workbook[worksheet].delete_rows(idx=index_row[row_del], amount=1)
            index_row = list(map(lambda k: k - 1, index_row))
            
    def sum_of_total_value(self, worksheet_name: str, column_to_sum: int):
        sum_of_total_value = 0
        
        for row in self.workbook[worksheet_name].iter_rows(min_row=2, values_only=True):
            cell_value = float(row[column_to_sum-1])
            if cell_value is not None:
                sum_of_total_value += cell_value
        return sum_of_total_value
    
    def last_row_index(self, worksheet_name, column_index: int):
        last_row_indx = None
        for row in self.workbook[worksheet_name].iter_rows(min_row=1, min_col=column_index, max_col=column_index):
            cell_value = row[0].value
            if cell_value is not None:
                last_row_indx = row[0].row  
        return last_row_indx

    def writing_total(self, worksheet_name, last_row_indx:int, column_indx:int, sum_of_values:float):
        if last_row_indx is not None:
            self.workbook[worksheet_name].cell(row=last_row_indx + 1, column= column_indx, value=sum_of_values)

    def apply_currency_format_columns(self, specific_worksheets, column_to_format:int):
        specific_worksheets = [self.workbook[worksheet_name] for worksheet_name in specific_worksheets]
        
        for worksheet in specific_worksheets:
            maxr = worksheet.max_row
            for cell in worksheet.iter_rows(min_row=2,max_row=maxr, min_col=column_to_format, max_col=column_to_format, values_only=False):
                cell[0].style = self.currency_format
    
    def center_values_rows(self, specific_worksheets, column_index:int):
        specific_worksheets = [self.workbook[worksheet_name] for worksheet_name in specific_worksheets]
        for worksheet in specific_worksheets:
            maxr = worksheet.max_row
            for cell in worksheet.iter_rows(min_row=2, max_row=maxr,min_col=column_index, max_col=column_index, values_only=False):
                cell[0].alignment = Alignment(horizontal='center')
    
    def center_values_columns(self, specific_worksheets, row_index:int):
        specific_worksheets = [self.workbook[worksheet_name] for worksheet_name in specific_worksheets]
        for worksheet in specific_worksheets:
            maxc = worksheet.max_column
            for cell in worksheet.iter_cols(min_col=1,max_col=maxc,min_row= row_index, max_row=row_index, values_only=False):
                cell[0].alignment = Alignment(horizontal='center')
                cell[0].font = Font(bold=True)

    def write_total_and_apply_percentage_format_one_cell(self, worksheet_name, row_to_format:int, column_to_format:int, total_value:float):
        worksheet = self.workbook[worksheet_name]
        worksheet.cell(row_to_format, column_to_format, total_value).style = self.percentage_format
    
    def write_total_and_apply_currency_format_one_cell(self, worksheet_name, row_to_format:int, column_to_format:int, total_value:float):
        worksheet = self.workbook[worksheet_name]
        worksheet.cell(row_to_format, column_to_format, total_value).style = self.currency_format

    def avg_each_total(self, value_perc:float, value2:float):
        total_sum = value_perc + value2
        percentage = value_perc/total_sum

        return percentage
        
    def save_file(self, new_sheet_name):
        file_name = "REPORTE MES " + new_sheet_name

        directory_path = os.path.dirname(self.file_path)

        new_file_path = directory_path + "\\" + file_name + ".xlsx"
        try:
            self.workbook.save(new_file_path)
            print("Archivo generado!")
        except PermissionError:
            print("No se logró guardar el archivo")       

# Usage
file_path = input("Escribe la dirección del archivo: ").strip(' "')
report = ExcelReport(file_path)
report.load_workbook()

# Changing the worksheets name to month's name
old_sheet_name = "CFDI's"
new_sheet_name = input("Escribe el mes del reporte: ").upper()
report.change_worksheet_name(old_sheet_name, new_sheet_name)

# Delete not used columns in the workbook
columns_to_delete = [1,3,4,5,6,9,11,12,14]
sheet_name = new_sheet_name
report.delete_columns(sheet_name, columns_to_delete)

# Delete rows that have 'CANCELADO' or 'PENDIENTE'
sheet_name = new_sheet_name
values_to_delete = ("CANCELADO", "PENDIENTE")
report.delete_rows(sheet_name,column_index=3, values_to_delete=values_to_delete)

# Select which invoices correspond to which office
report.open_invoices("C:/Users/alexi/Documents/Personal/Python Learning/ReportsCreator Project/facturas_juarez.txt", "C:/Users/alexi/Documents/Personal/Python Learning/ReportsCreator Project/facturas_chih.txt", "C:/Users/alexi/Documents/Personal/Python Learning/ReportsCreator Project/facturas_jrzychih.txt")

# Filter and append rows to the desired office
sheet_name = new_sheet_name
report.filter_and_append_rows(worksheet_name=sheet_name)

# Creating new worksheets
worksheet_name_jrz = 'JRZ'
worksheet_name_chih = 'CHIH'
worksheet_jrz = report.create_new_worksheet(worksheet_name_jrz)
worksheet_chih = report.create_new_worksheet(worksheet_name_chih)

# Change columns width
sheet_name = report.workbook[new_sheet_name]

columns_to_change = ["D", "E", "F", "I", "J"]

specific_worksheets = [new_sheet_name, 'JRZ', 'CHIH']
report.change_column_width(columns_to_change, 13, specific_worksheets)
report.change_column_width("B", 70, specific_worksheets)

# Copy rows to corresponding ws
rows_to_copy_jrz = report.rows_juarez
rows_to_copy_chih = report.rows_chih

report.copy_rows_to_ws(rows_to_copy_jrz, new_sheet_name, worksheet_name_jrz)
report.copy_rows_to_ws(rows_to_copy_chih, new_sheet_name, worksheet_name_chih)

# Deleting empty rows from each worksheet
report.delete_empty_rows(worksheet_name_jrz)
report.delete_empty_rows(worksheet_name_chih)

# Suming up total values
sum_values_chih = report.sum_of_total_value(worksheet_name_chih, 6)
sum_values_jrz = report.sum_of_total_value(worksheet_name_jrz, 6)

# Finding last row index
last_row_jrz = report.last_row_index(worksheet_name_jrz,6)
last_row_chih = report.last_row_index(worksheet_name_chih,6)

# Writing totals
report.writing_total(worksheet_name_jrz, last_row_jrz, 6, sum_values_jrz)
report.writing_total(worksheet_name_chih, last_row_chih, 6, sum_values_chih)

# Apply currency format
specific_worksheets = [new_sheet_name, 'JRZ', 'CHIH']
report.apply_currency_format_columns(specific_worksheets, 6)

# Center values
report.center_values_rows(specific_worksheets,2)
report.center_values_columns(specific_worksheets, 1)

# Writing avg
source_worksheet = report.workbook[new_sheet_name]
jrz_cell = source_worksheet['H6'] = "JRZ"
chih_cell = source_worksheet["H7"] = "CHIH"

avg_jrz = report.avg_each_total(sum_values_jrz,sum_values_chih)
avg_chih = report.avg_each_total(sum_values_chih, sum_values_jrz)

report.write_total_and_apply_currency_format_one_cell(new_sheet_name,6,9,sum_values_jrz)
source_worksheet.cell(6,9).font = Font(bold=True)
report.write_total_and_apply_currency_format_one_cell(new_sheet_name,7,9,sum_values_chih)
source_worksheet.cell(7,9).font = Font(bold=True)

report.write_total_and_apply_percentage_format_one_cell(new_sheet_name,6,10,avg_jrz)
source_worksheet.cell(6,10).font = Font(bold=True)
report.write_total_and_apply_percentage_format_one_cell(new_sheet_name,7,10,avg_chih)
source_worksheet.cell(7,10).font = Font(bold=True)

# Save file
report.save_file(new_sheet_name)