import os
import openpyxl
import time
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import NamedStyle, Font, Alignment

# Select the file path
try:
    file_path = input("Escribe la dirección del archivo: ").strip(' "')
    workbook = openpyxl.load_workbook(file_path)  # Load the report workbook

except openpyxl.utils.exceptions.InvalidFileException:
    print("Error: Invalid file format. Openpyxl does not support this file format.")
    exit()

# Changing the worksheets name to month's name
old_sheet_name = "CFDI's"
new_sheet_name = input("Escribe el mes del reporte: ").upper()

worksheet = workbook[old_sheet_name]
worksheet.title = new_sheet_name

# Deleting not used columns in the report
columns_to_delete = [1, 3, 4, 5, 6, 9, 11, 12, 14]

for col_index in sorted(columns_to_delete, reverse=True):
    worksheet.delete_cols(col_index)

# Delete all the rows that contain 'PENDIENTE' and 'CANCELADO'
rows_to_delete = []

for row in worksheet.iter_rows(min_row=2,min_col=3, max_col=3, values_only=False):
    cell_value = row[0].value
    if cell_value in ("CANCELADO", "PENDIENTE"):
        row_index = row[0].row
        rows_to_delete.append(row_index)

for row_index in sorted(rows_to_delete, reverse=True):
    worksheet.delete_rows(row_index)

# Select which invoices correspond to which office
try:
    with open("C:/Users/alexi/Documents/Personal/Python Learning/ReportsCreator Project/facturas_juarez.txt", "r") as file:
        names_to_search_jrz = [name.strip() for name in file.readlines()]
    with open("C:/Users/alexi/Documents/Personal/Python Learning/ReportsCreator Project/facturas_chih.txt", "r") as file:
        names_to_search_chih = [name.strip() for name in file.readlines()]
    with open("C:/Users/alexi/Documents/Personal/Python Learning/ReportsCreator Project/facturas_jrzychih.txt", "r") as file:
        names_to_search_both = [name.strip() for name in file.readlines()]
except FileNotFoundError:
    print("Error: Archivo de los nombres de las empresas no encontrado")
    exit()

# Adding the header to be copied
rows_juarez = [1]
rows_chih = [1]
row_not_in_count = []

max_row = worksheet.max_row - 2

try:
    for row in worksheet.iter_rows(min_row=2,max_row=max_row):
        cell_value = row[1].value
        cell_total = row[5].value
        if cell_value in names_to_search_jrz:
            row_index = row[0].row
            rows_juarez.append(row_index)
        elif cell_value in names_to_search_chih:
            row_index_cuu = row[0].row
            rows_chih.append(row_index_cuu)
        elif cell_value in names_to_search_both:
            row_index = row[0].row
            if cell_value == "NEFAB MEXICO":
                if cell_total >= 100000:
                    rows_juarez.append(row_index)
                else:
                    rows_chih.append(row_index)
            else:
                if 150000 >= cell_total >= 130000:
                    rows_juarez.append(row_index)
                elif cell_total < 30000:
                    row_not_in_count.append(row_index)
                else:
                    rows_chih.append(row_index)
        else:
            print(str(cell_value) + " no está en el archivo de los nombres de las empresas")
            exit()
except:
    print("Error: Archivo de los nombres de las empresas no encontrado")
    exit()

# Create a new worksheet to store the copied rows
workbook.create_sheet("JRZ")
workbook.create_sheet("CHIH")
worksheet_jrz = workbook["JRZ"]
worksheet_chih = workbook["CHIH"]

source_worksheet = workbook[new_sheet_name]

maxr = source_worksheet.max_row
maxc = source_worksheet.max_column

# Changing column width
columns_width_change_source = ["D", "E", "F", "I", "J"]

worksheets = [source_worksheet, worksheet_chih, worksheet_jrz]

for worksheet in worksheets:
    for column in columns_width_change_source:
        worksheet.column_dimensions[column].width = 13
    worksheet.column_dimensions["B"].width = 70

# Copy the rows that correspond to 'names_to_search' from the original worksheet to the new worksheet
for r in rows_juarez:
    for c in range (1, maxc + 1):
        worksheet_jrz.cell(row=r, column=c).value = source_worksheet.cell(row=r, column=c).value
for r in rows_chih:
    for c in range (1, maxc + 1):
        worksheet_chih.cell(row=r, column=c).value = source_worksheet.cell(row=r, column=c).value

# Deleting empty rows from "JRZ" sheet
index_row = []

for i in range(1,worksheet_jrz.max_row):
    if worksheet_jrz.cell(i,1).value is None:
        index_row.append(i)

for row_del in range(len(index_row)):
    worksheet_jrz.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k - 1, index_row))   

# Deleting empty rows from "CHIH" sheet
index_row = []

for i in range(1,worksheet_chih.max_row):
    if worksheet_chih.cell(i,1).value is None:
        index_row.append(i)

for row_del in range(len(index_row)):
    worksheet_chih.delete_rows(idx=index_row[row_del], amount=1)
    index_row = list(map(lambda k: k - 1, index_row))   

# Sum of total value
column_to_sum = 6
sum_of_values_jrz = 0
sum_of_values_chih = 0

for row in worksheet_jrz.iter_rows(min_row=2, values_only=True):
    cell_value = float(row[column_to_sum-1]) # Adjust the column index to 0-based indexing
    if cell_value is not None:
        sum_of_values_jrz += cell_value

for row in worksheet_chih.iter_rows(min_row=2, values_only=True):
    cell_value_chih = float(row[column_to_sum-1]) # Adjust the column index to 0-based indexing
    if cell_value_chih is not None:
        sum_of_values_chih += cell_value_chih

# Find last row index
last_row_index_jrz = None
for row in worksheet_jrz.iter_rows(min_row=1, min_col=6, max_col=6):
    cell_value = row[0].value
    if cell_value is not None:
        last_row_index_jrz = row[0].row

last_row_index_chih = None
for row in worksheet_chih.iter_rows(min_row=1, min_col=6, max_col=6):
    cell_value_chih = row[0].value
    if cell_value_chih is not None:
        last_row_index_chih = row[0].row

# Write Total
if last_row_index_jrz is not None:
    worksheet_jrz.cell(row=last_row_index_jrz + 1, column=6, value=sum_of_values_jrz)
if last_row_index_chih is not None:
    worksheet_chih.cell(row=last_row_index_chih + 1, column=6, value=sum_of_values_chih)

# Create named styles for currency and percentage
currency_format = NamedStyle(name='currency', number_format='"$"#,##0.00')
percentage_format = NamedStyle(name='percentage', number_format='0.00%')

# Apply currency formatting to the cell in each worksheet
for worksheet in worksheets:
    for cell in worksheet.iter_rows(min_row=2, max_row=maxr, min_col=6, max_col=6, values_only=False):
        cell[0].style =currency_format

# Center values in each worksheet
for worksheet in worksheets:
    for cell in worksheet.iter_rows(min_row=2, max_row=maxr, min_col=2, max_col=2, values_only=False):
        cell[0].alignment = Alignment(horizontal='center')
    for cell in worksheet.iter_cols(min_col=1, max_col=maxc,min_row=1, max_row=1, values_only=False):
        cell[0].alignment = Alignment(horizontal='center')
        cell[0].font = Font(bold=True)

# Writing the avg for each one
jrz_cell = source_worksheet['H6'] = "JRZ"
chih_cell = source_worksheet["H7"] = "CHIH"

# Write the total values and apply the currency format
source_worksheet.cell(6,9,sum_of_values_jrz).style = currency_format
source_worksheet.cell(7,9,sum_of_values_chih).style = currency_format

# Calculate the percentages
total_sum = sum_of_values_chih + sum_of_values_jrz
percentage__jrz = (sum_of_values_jrz) / (total_sum)
percentage__chih = (sum_of_values_chih / total_sum)

# Write the percentages and apply bold formatting
source_worksheet.cell(6,10,percentage__jrz).style = percentage_format
source_worksheet.cell(6,10, percentage__jrz).font = Font(bold=True)

source_worksheet.cell(7,10,percentage__chih).style = percentage_format
source_worksheet.cell(7,10, percentage__chih).font = Font(bold=True)

# Change file's name
file_name = "REPORTE MES " + new_sheet_name

# Saving the workbook
directory_path = os.path.dirname(file_path)

new_file_path = directory_path + "\\" + file_name + ".xlsx"
try:
    workbook.save(new_file_path)
    print("Archivo generado!")
except PermissionError:
    print("No se logró guardar el archivo")
